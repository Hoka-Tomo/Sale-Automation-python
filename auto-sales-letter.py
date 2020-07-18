# PythonでExcelを操作できるライブラリをインポート
import openpyxl

# セルに書き込み
# Excelファイルの読み込み
book = openpyxl.load_workbook("sales_list.xlsx")

# シート1の取得
sheet1 = book["Sheet1"]
## 取得行を変える際は、ここを変更する。
##（本来は自動化をすすめるべきところだが、いったん。取得の際に以下の情報をSheet1から取得するが、いちいち変えるのが面倒なため。）
rownum = 4

sheet1contact = sheet1.cell(row=rownum,column=10)
sheet1title = sheet1.cell(row=rownum,column=3)
sheet1url = sheet1.cell(row=rownum,column=4)
sheet1cms = sheet1.cell(row=rownum,column=5)
sheet1checkpage = sheet1.cell(row=rownum,column=7)
sheet1score = sheet1.cell(row=rownum,column=8)

# シート3の取得
sheet3 = book["Sheet3"]

# シートの指定セルに書き込み（固定）
## 連絡先
sheet3.cell(row=1,column=2).value = sheet1contact.value
## サイト名
sheet3.cell(row=4,column=2).value = sheet1title.value
sheet3.cell(row=6,column=2).value = sheet1title.value
sheet3.cell(row=11,column=3).value = sheet1title.value
sheet3.cell(row=21,column=2).value = sheet1title.value
## CMS
sheet3.cell(row=11,column=5).value = sheet1cms.value
## URL
sheet3.cell(row=21,column=4).value = sheet1url.value
## 対象ページ
sheet3.cell(row=21,column=7).value = sheet1checkpage.value
## スコア
sheet3.cell(row=21,column=8).value = sheet1score.value

# 保存
book.save("sales_list.xlsx")